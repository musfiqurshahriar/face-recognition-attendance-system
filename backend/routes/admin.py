from flask import Blueprint, render_template, request, send_file, redirect, url_for, flash
from flask_login import login_required, current_user
from database import SessionLocal, Attendance, load_students_from_excel, load_teachers_from_excel, get_admin_from_env
from sqlalchemy import func
from datetime import datetime
import pandas as pd
import io
import os
import time

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

_DASHBOARD_CACHE = {}
_DASHBOARD_CACHE_TIME = 0
CACHE_TTL = 60

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ["admin", "teacher"]:
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)
    return decorated

DESIGNATION_RANK = {
    "Professor": 1,
    "Associate Professor": 2,
    "Assistant Professor": 3,
    "Lecturer": 4
}

def get_teacher_rank(designation):
    return DESIGNATION_RANK.get(designation, 99)

def clear_dashboard_cache():
    global _DASHBOARD_CACHE, _DASHBOARD_CACHE_TIME
    _DASHBOARD_CACHE = {}
    _DASHBOARD_CACHE_TIME = 0

@admin_bp.route("/dashboard")
@login_required
@admin_required
def dashboard():
    global _DASHBOARD_CACHE, _DASHBOARD_CACHE_TIME

    today = __import__("datetime").date.today().strftime("%Y-%m-%d")
    now = time.time()

    if _DASHBOARD_CACHE and (now - _DASHBOARD_CACHE_TIME) < CACHE_TTL and _DASHBOARD_CACHE.get("today") == today:
        return render_template("admin/dashboard.html", **_DASHBOARD_CACHE)

    db = SessionLocal()
    try:
        all_today_records = db.query(Attendance).filter(
            Attendance.date == today
        ).all()

        def student_obj_sort_key(r):
            sec = r.section if r.section else "0-0"
            try:
                first_year = int(sec.split("-")[0])
            except:
                first_year = 0
            return (-first_year, r.roll_number if r.roll_number else "")

        student_records = sorted(
            [r for r in all_today_records if r.role == "student"],
            key=student_obj_sort_key
        )

        teacher_records = sorted(
            [r for r in all_today_records if r.role == "teacher"],
            key=lambda x: get_teacher_rank(x.section)
        )

        today_records = student_records + teacher_records

        students_list = load_students_from_excel()
        total_students = len(students_list)
        today_present = len(student_records)
        today_absent = total_students - today_present
        sections = sorted(list(set([s["section"] for s in students_list if s.get("section")])))

        def student_dict_sort_key(s):
            sec = s.get("section", "0-0")
            try:
                first_year = int(sec.split("-")[0])
            except:
                first_year = 0
            return (-first_year, s.get("roll", ""))

        present_rolls = [r.roll_number for r in student_records]
        absent_students = sorted(
            [s for s in students_list if s.get("roll") not in present_rolls],
            key=student_dict_sort_key
        )

        from database import load_teachers_from_excel
        teachers_list = load_teachers_from_excel()
        present_teacher_names = [r.name for r in teacher_records]
        absent_teachers = sorted(
            [t for t in teachers_list if t.get("name") not in present_teacher_names],
            key=lambda x: get_teacher_rank(x.get("designation", ""))
        )

        all_dates = db.query(Attendance.date).filter(Attendance.role == "student").distinct().all()
        total_days = len(all_dates)

        low_attendance = []
        if total_days > 0:
            attendance_results = db.query(Attendance.roll_number, func.count(Attendance.id))\
                .filter(Attendance.role == "student")\
                .group_by(Attendance.roll_number).all()

            attendance_counts = {row[0]: row[1] for row in attendance_results}

            for student in students_list:
                student_roll = student.get("roll")
                present_count = attendance_counts.get(student_roll, 0)
                percentage = round((present_count / total_days * 100), 1)
                if percentage < 70:
                    low_attendance.append({
                        "roll": student_roll,
                        "name": student.get("name"),
                        "section": student.get("section"),
                        "percentage": percentage,
                        "present": present_count,
                        "total": total_days
                    })

        from datetime import datetime as dt, timedelta
        chart_labels = []
        chart_present = []
        chart_absent = []

        start_date = (dt.today() - timedelta(days=6)).strftime("%Y-%m-%d")

        chart_results = db.query(Attendance.date, func.count(Attendance.roll_number.distinct()))\
            .filter(Attendance.date >= start_date, Attendance.role == "student")\
            .group_by(Attendance.date).all()

        chart_counts = {row[0]: row[1] for row in chart_results}

        for i in range(6, -1, -1):
            day = (dt.today() - timedelta(days=i)).strftime("%Y-%m-%d")
            day_label = (dt.today() - timedelta(days=i)).strftime("%d %b")
            present = chart_counts.get(day, 0)
            chart_labels.append(day_label)
            chart_present.append(present)
            chart_absent.append(total_students - present)

        context = dict(
            today_records=today_records,
            total_students=total_students,
            today_present=today_present,
            today_absent=today_absent,
            sections=sections,
            today=today,
            absent_students=absent_students,
            absent_teachers=absent_teachers,
            low_attendance=low_attendance,
            chart_labels=chart_labels,
            chart_present=chart_present,
            chart_absent=chart_absent,
            all_registered_students=students_list,
            all_registered_teachers=teachers_list
        )

        _DASHBOARD_CACHE.clear()
        _DASHBOARD_CACHE.update(context)
        _DASHBOARD_CACHE_TIME = now

    finally:
        db.close()

    return render_template("admin/dashboard.html", **_DASHBOARD_CACHE)


@admin_bp.route("/attendance")
@login_required
@admin_required
def attendance():
    db = SessionLocal()
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    section = request.args.get("section", "")
    semester = request.args.get("semester", "")

    # Default: শুধু শেষ ৭ দিনের data দেখাবে
    if not date_from and not date_to:
        from datetime import timedelta
        date_from = (datetime.today() - timedelta(days=7)).strftime("%Y-%m-%d")

    s_query = db.query(Attendance).filter(Attendance.role == "student")
    t_query = db.query(Attendance).filter(Attendance.role == "teacher")

    if date_from:
        s_query = s_query.filter(Attendance.date >= date_from)
        t_query = t_query.filter(Attendance.date >= date_from)
    if date_to:
        s_query = s_query.filter(Attendance.date <= date_to)
        t_query = t_query.filter(Attendance.date <= date_to)
    if section:
        s_query = s_query.filter(Attendance.section == section)
    if semester:
        s_query = s_query.filter(Attendance.semester == semester)

    student_records = s_query.order_by(
        Attendance.date.desc(),
        Attendance.section.desc(),
        Attendance.roll_number
    ).limit(500).all()  # ← সর্বোচ্চ ৫০০ record

    all_teacher_records = t_query.limit(100).all()  # ← সর্বোচ্চ ১০০ record
    teacher_records = sorted(
        all_teacher_records,
        key=lambda x: (x.date, get_teacher_rank(x.section))
    )

    students_list = load_students_from_excel()
    sections = sorted(list(set([s["section"] for s in students_list if s.get("section")])))
    semesters = db.query(Attendance.semester).distinct().all()
    semesters = [s[0] for s in semesters if s[0]]

    db.close()
    return render_template("admin/attendance.html",
        student_records=student_records,
        teacher_records=teacher_records,
        sections=sections,
        semesters=semesters,
        date_from=date_from,
        date_to=date_to,
        selected_section=section,
        selected_semester=semester
    )


@admin_bp.route("/percentage")
@login_required
@admin_required
def percentage():
    db = SessionLocal()
    section = request.args.get("section", "")
    semester = request.args.get("semester", "")

    all_students = load_students_from_excel()
    if section:
        students = [s for s in all_students if s.get("section") == section]
    else:
        students = all_students

    def session_sort_key(s):
        session = s.get("section", "0-0")
        try:
            first_year = int(session.split("-")[0])
        except:
            first_year = 0
        return (-first_year, s.get("roll", ""))

    students = sorted(students, key=session_sort_key)

    total_days = db.query(Attendance.date).filter(
        Attendance.role == "student"
    ).distinct().count()

    result = []
    for student in students:
        student_roll = student.get("roll")
        count_query = db.query(Attendance).filter(
            Attendance.role == "student",
            Attendance.roll_number == student_roll
        )
        if semester:
            count_query = count_query.filter(Attendance.semester == semester)
        present_count = count_query.count()
        percentage = round((present_count / total_days * 100), 1) if total_days > 0 else 0
        result.append({
            "roll": student_roll,
            "name": student.get("name"),
            "section": student.get("section"),
            "present": present_count,
            "total": total_days,
            "percentage": percentage
        })

    sections = sorted(list(set([s["section"] for s in all_students if s.get("section")])))
    semesters = db.query(Attendance.semester).distinct().all()
    semesters = [s[0] for s in semesters if s[0]]

    db.close()
    return render_template("admin/percentage.html",
        result=result,
        sections=sections,
        semesters=semesters,
        selected_section=section,
        selected_semester=semester
    )


@admin_bp.route("/export/excel")
@login_required
@admin_required
def export_excel():
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    db = SessionLocal()
    section = request.args.get("section", "")
    semester = request.args.get("semester", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    export_type = request.args.get("type", "student")

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wb = writer.book

        if export_type == "student" or export_type == "all":
            s_query = db.query(Attendance).filter(Attendance.role == "student")
            if section:
                s_query = s_query.filter(Attendance.section == section)
            if semester:
                s_query = s_query.filter(Attendance.semester == semester)
            if date_from:
                s_query = s_query.filter(Attendance.date >= date_from)
            if date_to:
                s_query = s_query.filter(Attendance.date <= date_to)

            s_records = s_query.order_by(
                Attendance.date,
                Attendance.section.desc(),
                Attendance.roll_number
            ).all()

            ws = wb.create_sheet("Students")
            headers = ["Date", "Roll", "Name", "Session", "Time", "Status", "Semester"]
            ws.append(headers)

            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font

            current_date = None
            row_num = 2

            for r in s_records:
                if r.date != current_date:
                    current_date = r.date
                    date_cell = ws.cell(row=row_num, column=1, value=r.date)
                    date_cell.font = Font(bold=True, size=12, color="000000")
                    date_cell.fill = PatternFill(start_color="E8E4FF", end_color="E8E4FF", fill_type="solid")
                    ws.row_dimensions[row_num].height = 20
                    row_num += 1

                ws.cell(row=row_num, column=1, value="")
                ws.cell(row=row_num, column=2, value=r.roll_number)
                ws.cell(row=row_num, column=3, value=r.name)
                ws.cell(row=row_num, column=4, value=r.section)
                ws.cell(row=row_num, column=5, value=r.time)
                ws.cell(row=row_num, column=6, value=r.status)
                ws.cell(row=row_num, column=7, value=r.semester)
                row_num += 1

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

        if export_type == "teacher" or export_type == "all":
            t_query = db.query(Attendance).filter(Attendance.role == "teacher")
            if date_from:
                t_query = t_query.filter(Attendance.date >= date_from)
            if date_to:
                t_query = t_query.filter(Attendance.date <= date_to)

            all_t_records = t_query.all()
            t_records = sorted(
                all_t_records,
                key=lambda x: (x.date, get_teacher_rank(x.section))
            )

            t_ws = wb.create_sheet("Teachers")
            t_headers = ["Date", "Name", "Designation", "Time", "Status"]
            t_ws.append(t_headers)

            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col in range(1, len(t_headers) + 1):
                cell = t_ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font

            current_date_t = None
            row_num_t = 2

            for r in t_records:
                if r.date != current_date_t:
                    current_date_t = r.date
                    date_cell = t_ws.cell(row=row_num_t, column=1, value=r.date)
                    date_cell.font = Font(bold=True, size=12, color="000000")
                    date_cell.fill = PatternFill(start_color="E8E4FF", end_color="E8E4FF", fill_type="solid")
                    t_ws.row_dimensions[row_num_t].height = 20
                    row_num_t += 1

                t_ws.cell(row=row_num_t, column=1, value="")
                t_ws.cell(row=row_num_t, column=2, value=r.name)
                t_ws.cell(row=row_num_t, column=3, value=r.section)
                t_ws.cell(row=row_num_t, column=4, value=r.time)
                t_ws.cell(row=row_num_t, column=5, value=r.status)
                row_num_t += 1

            for col in range(1, len(t_headers) + 1):
                t_ws.column_dimensions[get_column_letter(col)].width = 20

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    db.close()
    output.seek(0)

    filename = f"{export_type}_attendance.xlsx" if export_type != "all" else "full_attendance_report.xlsx"

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@admin_bp.route("/send-notifications", methods=["POST"])
@login_required
@admin_required
def send_notifications():
    from email_sender import send_absent_notifications
    target_date = request.form.get("date", "")
    if not target_date:
        from datetime import date
        target_date = date.today().strftime("%Y-%m-%d")

    result = send_absent_notifications(target_date)
    flash(
        f"{target_date} তারিখের অনুপস্থিত {result['total_absent']} জনের মধ্যে "
        f"{result['success']} জনকে email পাঠানো হয়েছে।",
        "success"
    )
    return redirect(url_for("admin.dashboard"))


@admin_bp.route("/upload-students", methods=["GET", "POST"])
@login_required
@admin_required
def upload_students():
    if request.method == "POST":
        file = request.files.get("excel_file")
        if file and file.filename.endswith(".xlsx"):
            save_path = os.path.normpath(os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                "..", "..", "database", "students.xlsx"
            ))
            file.save(save_path)
            flash("Student list সফলভাবে update হয়েছে!", "success")
            return redirect(url_for("admin.dashboard"))
        else:
            flash("শুধু .xlsx ফাইল upload করুন!", "error")

    return render_template("admin/upload_students.html")


@admin_bp.route("/change-password", methods=["GET", "POST"])
@login_required
@admin_required
def change_password():
    if request.method == "POST":
        current = request.form.get("current_password", "")
        new_pass = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")

        admin = get_admin_from_env()

        if current != admin["login_password"]:
            flash("বর্তমান password ভুল!", "error")
            return redirect(url_for("admin.change_password"))

        if new_pass != confirm:
            flash("নতুন password দুটো মিলছে না!", "error")
            return redirect(url_for("admin.change_password"))

        if len(new_pass) < 6:
            flash("Password কমপক্ষে ৬ অক্ষরের হতে হবে!", "error")
            return redirect(url_for("admin.change_password"))

        env_path = os.path.normpath(os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "..", ".env"
        ))

        with open(env_path, "r") as f:
            lines = f.readlines()

        with open(env_path, "w") as f:
            for line in lines:
                if line.startswith("ADMIN_PASSWORD="):
                    f.write(f"ADMIN_PASSWORD={new_pass}\n")
                else:
                    f.write(line)

        from dotenv import load_dotenv
        load_dotenv(env_path, override=True)

        flash("Password সফলভাবে পরিবর্তন হয়েছে!", "success")
        return redirect(url_for("admin.dashboard"))

    return render_template("admin/change_password.html")


@admin_bp.route("/export/percentage")
@login_required
@admin_required
def export_percentage():
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    db = SessionLocal()
    section = request.args.get("section", "")
    semester = request.args.get("semester", "")

    all_students = load_students_from_excel()
    if section:
        students = [s for s in all_students if s.get("section") == section]
    else:
        students = all_students

    def session_sort_key(s):
        session = s.get("section", "0-0")
        try:
            first_year = int(session.split("-")[0])
        except:
            first_year = 0
        return (-first_year, s.get("roll", ""))

    students = sorted(students, key=session_sort_key)

    total_days = db.query(Attendance.date).filter(
        Attendance.role == "student"
    ).distinct().count()

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wb = writer.book
        ws = wb.create_sheet("Percentage")

        headers = ["Roll", "Name", "Session", "Present", "Total Class", "Percentage"]
        ws.append(headers)

        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        for student in students:
            student_roll = student.get("roll")
            count_query = db.query(Attendance).filter(
                Attendance.role == "student",
                Attendance.roll_number == student_roll
            )
            if semester:
                count_query = count_query.filter(Attendance.semester == semester)
            present_count = count_query.count()
            percentage = round((present_count / total_days * 100), 1) if total_days > 0 else 0

            row = [
                student_roll,
                student.get("name"),
                student.get("section"),
                present_count,
                total_days,
                f"{percentage}%"
            ]
            ws.append(row)

            if percentage < 70:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = PatternFill(
                        start_color="FDECEA", end_color="FDECEA", fill_type="solid"
                    )

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    db.close()
    output.seek(0)
    return send_file(
        output,
        download_name="attendance_percentage.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@admin_bp.route("/manual-attendance", methods=["POST"])
@login_required
@admin_required
def manual_attendance():
    role = request.form.get("role")
    identifier = request.form.get("identifier", "").strip()
    status = request.form.get("status", "On Time")
    custom_date = request.form.get("date")

    if not custom_date:
        custom_date = datetime.now().strftime("%Y-%m-%d")

    time_str = datetime.now().strftime("%I:%M %p")
    db = SessionLocal()

    try:
        if role == "student" and identifier.lower() == "all":
            all_students = load_students_from_excel()
            count = 0
            for student in all_students:
                exists = db.query(Attendance).filter(
                    Attendance.name == student["name"],
                    Attendance.date == custom_date
                ).first()
                if not exists:
                    new_record = Attendance(
                        user_id=student.get("roll", student["name"]),
                        name=student["name"],
                        role="student",
                        roll_number=student.get("roll", ""),
                        section=student.get("section", ""),
                        date=custom_date,
                        time=time_str,
                        status=status,
                        semester=student.get("semester", "")
                    )
                    db.add(new_record)
                    count += 1
            db.commit()
            clear_dashboard_cache()
            flash(f"সফলভাবে মোট {count} জন শিক্ষার্থীর বাল্ক হাজিরা নেওয়া হয়েছে।", "success")

        else:
            target_name = identifier
            target_roll = ""
            target_section = ""
            target_semester = ""
            db_role = "student" if role == "student" else "teacher"

            if role == "student":
                all_students = load_students_from_excel()
                student_match = next(
                    (s for s in all_students if str(s.get("roll")) == identifier), None
                )
                if student_match:
                    target_name = student_match["name"]
                    target_roll = student_match.get("roll", "")
                    target_section = student_match.get("section", "")
                    target_semester = student_match.get("semester", "")
            else:
                all_teachers = load_teachers_from_excel()
                teacher_match = next(
                    (t for t in all_teachers if t["name"] == identifier), None
                )
                if teacher_match:
                    target_name = teacher_match["name"]
                    target_section = teacher_match.get("designation", "")

            exists = db.query(Attendance).filter(
                Attendance.name == target_name,
                Attendance.date == custom_date
            ).first()

            if exists:
                flash(f"দুঃখিত, {target_name} এর হাজিরা আজ আগেই নেওয়া হয়েছে!", "danger")
            else:
                new_record = Attendance(
                    user_id=target_roll if target_roll else target_name,
                    name=target_name,
                    role=db_role,
                    roll_number=target_roll,
                    section=target_section,
                    date=custom_date,
                    time=time_str,
                    status=status,
                    semester=target_semester
                )
                db.add(new_record)
                db.commit()
                clear_dashboard_cache()
                flash(f"সফলভাবে {target_name} এর ম্যানুয়াল হাজিরা নেওয়া হয়েছে।", "success")

    except Exception as e:
        db.rollback()
        print(f"Error: {e}")
    finally:
        db.close()

    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/update_attendance', methods=['POST'])
def update_attendance():
    date_to_update = request.form.get('update_date')
    role_type = request.form.get('role')
    identifier = request.form.get('student_roll')
    new_status = request.form.get('new_status')

    if date_to_update and identifier and new_status:
        session = SessionLocal()
        try:
            if role_type == 'student':
                record = session.query(Attendance).filter_by(date=date_to_update, roll_number=identifier).first()
            else:
                record = session.query(Attendance).filter_by(date=date_to_update, name=identifier).first()

            if record:
                if new_status == 'Absent':
                    session.delete(record)
                    session.commit()
                    clear_dashboard_cache()
                    flash(f"রেকর্ড মুছে ফেলা হয়েছে! এখন তাকে Absent লিস্টে দেখাবে।", "success")
                else:
                    record.status = new_status
                    session.commit()
                    clear_dashboard_cache()
                    flash(f"হাজিরা সফলভাবে '{new_status}' করা হয়েছে!", "success")
            else:
                flash(f"{date_to_update} তারিখে এই ব্যক্তির কোনো হাজিরার রেকর্ড পাওয়া যায়নি।", "warning")

        except Exception as e:
            session.rollback()
            flash("স্ট্যাটাস আপডেট করতে গিয়ে একটি সমস্যা হয়েছে।", "danger")
            print(f"[ERROR] {e}")
        finally:
            session.close()
    else:
        flash("অনুগ্রহ করে তারিখ, ব্যক্তি এবং নতুন স্ট্যাটাস নির্বাচন করুন।", "danger")

    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/delete_attendance', methods=['POST'])
def delete_attendance():
    date_to_delete = request.form.get('delete_date')
    delete_scope = request.form.get('delete_scope')

    if date_to_delete:
        session = SessionLocal()
        try:
            query = session.query(Attendance).filter_by(date=date_to_delete)

            if delete_scope == 'student':
                records_to_delete = query.filter_by(role='student').all()
            elif delete_scope == 'teacher':
                records_to_delete = query.filter_by(role='teacher').all()
            else:
                records_to_delete = query.all()

            if records_to_delete:
                for record in records_to_delete:
                    session.delete(record)
                session.commit()
                clear_dashboard_cache()
                flash(f"{date_to_delete} তারিখের সিলেক্টেড রেকর্ড সফলভাবে মুছে ফেলা হয়েছে!", "success")
            else:
                flash(f"{date_to_delete} তারিখে কোনো হাজিরার রেকর্ড পাওয়া যায়নি।", "warning")

        except Exception as e:
            session.rollback()
            flash("রেকর্ড মুছে ফেলতে গিয়ে একটি সমস্যা হয়েছে।", "danger")
            print(f"[ERROR] {e}")
        finally:
            session.close()
    else:
        flash("অনুগ্রহ করে একটি তারিখ নির্বাচন করুন।", "danger")

    return redirect(url_for('admin.dashboard'))